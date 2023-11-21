import mysql.connector
import openpyxl
from scraping_patriarca import *

def conectar_bbdd():
    configuracion = {
        'user': 'root',
        'password': '1234',
        'host': 'localhost',
        'database': 'patriarcadb',
        'port':'3307'
    }

    conexion= mysql.connector.connect(**configuracion)
    return conexion

def cerrar_conexion(conexion):
    conexion.commit()
    conexion.close()


def insertar_datos():

    #Obtenemos los datos del scraping
    list_mantecados = extraer_informacion()

    #Obtenemos conexion con la bbdd
    conexion = conectar_bbdd()

    #crear cursor
    cursor = conexion.cursor()

    script = "insert into mantecados (url,nombre,precio,descripcion) values (%s,%s,%s,%s)"

    for mantecados in list_mantecados:
        cursor.execute(script,(mantecados["url"],mantecados["nombre"], mantecados["precio"],mantecados["descripcion"]))

    cerrar_conexion(conexion)


def insertar_dato(nuevo_mantecado):

    #Obtenemos conexion con la bbdd
    conexion = conectar_bbdd()

    #crear cursor
    cursor = conexion.cursor()

    script = "insert into mantecados (url,nombre,precio,descripcion) values (%s,%s,%s,%s)"

    cursor.execute(script,(nuevo_mantecado["url"],nuevo_mantecado["nombre"], nuevo_mantecado["precio"],nuevo_mantecado["descripcion"]))

    cerrar_conexion(conexion)




def extraer_datos():

    #creamos lista
    lista_mantecados = []

    #abrir conexion
    conexion = conectar_bbdd()

    #abrimos cursor
    cursor = conexion.cursor(dictionary=True)

    #montamos el script
    script = "select * from mantecados"

    #ejecutamos la consulta
    cursor.execute(script)

    #Obtenemos columnas
    columnas = cursor.column_names

    #obtener los datos de la consulta
    lista_mantecados = cursor.fetchall()

    #cerramos conexion
    cerrar_conexion(conexion)

    return lista_mantecados,columnas



def crear_excel_con_datos():

    #Datos
    lista_mantecados,nombres_columnas = extraer_datos()

    #Creamos Hoja excel
    documento = openpyxl.Workbook()
    hoja_1 = documento.active

    #rellenamos la primera fila con los titulos
    for contador, nombre_columna in enumerate(nombres_columnas, start=1):
        hoja_1.cell(row=1,column=contador,value=nombre_columna)


    for fila,mantecado in enumerate(lista_mantecados, start=2):
        for columna,campo_mantecado in enumerate(mantecado.values(),start=1):
            hoja_1.cell(row=fila, column=columna,value=campo_mantecado)


    documento.save("mantecados.xlsx")


